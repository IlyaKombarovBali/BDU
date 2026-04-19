#!/usr/bin/env python3
"""
Загрузка каталога документов из первого листа xlsx в site.db (таблица doc).
Колонки файла: group, title, description, link → те же поля в БД (кроме id).

  python scripts/import_doc_xlsx.py
"""
from __future__ import annotations

import sqlite3
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import openpyxl  # noqa: E402

from auth import SITE_DB, upgrade_site_db_auth  # noqa: E402

XLSX = ROOT / "doc" / "Список документации по информационной безопасности предприятия.xlsx"


def _s(v) -> str | None:
    if v is None:
        return None
    t = str(v).strip()
    return t if t else None


def main() -> None:
    if not XLSX.is_file():
        raise SystemExit(f"Файл не найден: {XLSX}")

    upgrade_site_db_auth(SITE_DB)

    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not rows:
        raise SystemExit("Пустой первый лист")

    header = [str(x).strip().lower() if x is not None else "" for x in rows[0]]
    if header[:4] != ["group", "title", "description", "link"]:
        raise SystemExit(f"Неожиданный заголовок первого листа: {rows[0]!r}")

    out: list[tuple[str | None, str | None, str | None, str | None]] = []
    for r in rows[1:]:
        if not r or len(r) < 4:
            continue
        g, title, desc, link = r[0], r[1], r[2], r[3]
        if g is None and title is None:
            continue
        out.append((_s(g), _s(title), _s(desc), _s(link)))

    con = sqlite3.connect(str(SITE_DB))
    try:
        con.execute("DELETE FROM doc")
        con.executemany(
            'INSERT INTO doc ("group", title, description, link) VALUES (?,?,?,?)',
            out,
        )
        con.commit()
        n = con.execute("SELECT COUNT(*) FROM doc").fetchone()[0]
        print(f"Записей в doc: {n}")
    finally:
        con.close()


if __name__ == "__main__":
    main()
